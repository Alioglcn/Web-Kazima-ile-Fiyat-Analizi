import pandas as pd
import re
import openpyxl
from openpyxl.styles import PatternFill

# Fiyat sütunlarını güncelleyen fonksiyon
def update_price_column(df):
    for column in df.columns:
        if 'fiyat' in column:
            df[column] = df[column].apply(clean_convert_numeric)    
    return df

# Fiyat sütunlarını string ifadelerden temizler numeric formata çevirir
def clean_convert_numeric(price):
    if pd.notna(price):
        clean_price = re.sub(r'[^\d,]', '', str(price))
        if clean_price:
            return float(clean_price.replace(',', '.'))
    return price

# Excel dosyasından etiketleri okuyan fonksiyon
def read_keywords(file_path):
    # Excel dosyasını oku ve DataFrame'e dönüştür
    keywords_df = pd.read_excel(file_path)
    keywords_dict = {}

    # Her bir sütun için etiketleri bir sözlüğe ekleyerek sakla
    for column in keywords_df.columns:
        keywords_dict[column] = keywords_df[column].dropna().tolist()
    return keywords_dict

# Benzerlik skorlarını güncelleyen fonksiyon
def update_similarity_scores(df, keywords_dict, upgrade_dict):
    # Her rakip firma için etiketlere ve skor eşiklerine eriş
    for competitor, keywords in keywords_dict.items():
        upgrade_value = upgrade_dict.get(competitor, 0)
        
        # Her etiket için eşleşme kontrolü yap
        for keyword in keywords:
            keyword_lower = keyword.lower()
            
            # Veri setinde her bir satırı kontrol et
            for index, row in df.iterrows():
                eb_product = str(row['ebebek']).lower() if not pd.isna(row['ebebek']) else ''
                competitor_product = str(row[competitor]).lower() if competitor in df.columns and not pd.isna(row[competitor]) else ''

                # Eşleşme durumuna göre skoru düşür
                if keyword_lower in eb_product:
                    if keyword_lower not in competitor_product:
                        df.at[index, f'similarityscore{competitor}'] -= upgrade_value
                elif keyword_lower not in eb_product:
                    if keyword_lower in competitor_product:
                        df.at[index, f'similarityscore{competitor}'] -= upgrade_value
                        
                    

# Benzerlik skorlarına göre veri setini filtreleyen fonksiyon
def filter_dataframe(df, threshold_dict):
    # Her rakip firma için eşik skorlara eriş
    for competitor, threshold in threshold_dict.items():
        score_column = f'similarityscore{competitor}'
        mask = df[score_column] < threshold
        columns_to_reset = [competitor, f'{competitor}fiyat', score_column]

        # Eşik skoru altındaki satırları sıfırla
        for col in columns_to_reset:
            df.loc[mask, col] = ''

# Hücre renklerini kopyalayan fonksiyon
def copy_cell_color(source_cell, target_cell):
    # Kaynak hücrenin renk bilgilerini al
    if source_cell.fill.start_color.type == 'rgb':
        start_color_rgb = source_cell.fill.start_color.rgb
    else:
        start_color_rgb = None

    if source_cell.fill.end_color and source_cell.fill.end_color.type == 'rgb':
        end_color_rgb = source_cell.fill.end_color.rgb
    else:
        end_color_rgb = None

    # Hedef hücreye renk bilgilerini uygula
    target_cell.fill = PatternFill(start_color=start_color_rgb, end_color=end_color_rgb, fill_type=source_cell.fill.fill_type)

# Excel dosyalarındaki renkleri kopyalayan fonksiyon
def copy_excel_colors(source_file, target_file):
    # Kaynak ve hedef Excel dosyalarını yükle
    source_wb = openpyxl.load_workbook(source_file)
    target_wb = openpyxl.load_workbook(target_file)

    source_ws = source_wb.active
    target_ws = target_wb.active

    # Her hücrenin rengini kopyala
    for row in source_ws.iter_rows():
        for cell in row:
            target_cell = target_ws[cell.coordinate]
            copy_cell_color(cell, target_cell)

    # Hedef dosyayı kaydet ve kapat
    target_wb.save(target_file)
    source_wb.close()
    target_wb.close()

# Ana işlemi gerçekleştiren fonksiyon
def main():
    # Veri setini oku
    excel_file = 'beslenme.xlsx'
    df = pd.read_excel(excel_file)
    
    # Fiyat sütunlarını temizle ve nümerik yap
    df = update_price_column(df)

    # Etiketleri oku
    keywords_file = 'beslenme_labels.xlsx'
    keywords_dict = read_keywords(keywords_file)

    # Her rakip firma için eşik skorları belirle
    threshold_dict = {
        'civil': 0.74,
        'welcomebaby': 0.72,
        'babymall': 0.72,
        'madrenino': 0.73,
        'bebekhouse': 0.67
    }

    # Her rakip firma için benzerlik skorlarını güncelleyecek değerleri belirle
    upgrade_dict = {
        'civil': 0.16,
        'welcomebaby': 0.18,
        'babymall': 0.07,
        'madrenino': 0.01,
        'bebekhouse': 0.20
    }

    # Benzerlik skorlarını güncelle
    update_similarity_scores(df, keywords_dict, upgrade_dict)
    
    # Güncel skorlu eşleşme veri setini oluştur ve çıkar
    raw_data_file = 'beslenme_update_no_elimination.xlsx'
    df.to_excel(raw_data_file, index=False)
    
    # Benzerlik skorlarına göre filtrele
    filter_dataframe(df, threshold_dict)

    # Sonuçları Excel dosyasına yaz
    output_excel = 'beslenme_updated.xlsx'
    df.to_excel(output_excel, index=False)

    # Renkleri kopyala
    copy_excel_colors(excel_file, raw_data_file)
    copy_excel_colors(excel_file, output_excel)

if __name__ == "__main__":
    main()